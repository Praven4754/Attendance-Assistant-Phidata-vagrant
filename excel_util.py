import openpyxl
import os
import pandas as pd
from datetime import datetime, timedelta
import calendar
EXCEL_FILE = "attendance.xlsx"

def prefill_month(month: int, year: int):
    num_days = calendar.monthrange(year, month)[1]
    dates = [datetime(year, month, day) for day in range(1, num_days + 1)]

    data = {
        "Date": [d.strftime("%Y-%m-%d") for d in dates],
        "Day": [calendar.day_name[d.weekday()] for d in dates],
        "Status": ["Week Off" if d.weekday() in [5, 6] else "" for d in dates],  # 5=Saturday, 6=Sunday
        "Remarks": ["" for _ in dates],
    }

    df = pd.DataFrame(data)

    try:
        df_existing = pd.read_excel("attendance.xlsx")
        df_combined = pd.concat([df_existing, df]).drop_duplicates(subset="Date", keep="first")
    except FileNotFoundError:
        df_combined = df

    df_combined = df_combined.sort_values(by="Date")
    df_combined.to_excel("attendance.xlsx", index=False)
    return f"✅ Pre-filled {calendar.month_name[month]} {year} into attendance.xlsx"

import pandas as pd
def clear_entry(date: str):
    df = pd.read_excel("attendance.xlsx")
    df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%Y-%m-%d")

    if date not in df['Date'].values:
        raise ValueError(f"No entry found for {date} to clear.")
    
    df.loc[df['Date'] == date, ['Status', 'Remarks']] = ""
    df.to_excel("attendance.xlsx", index=False)

from tabulate import tabulate
# import pandas as pd
def fetch_monthly_timesheet(month_name: str) -> str:
    try:
        df = pd.read_excel("attendance.xlsx")
        df.fillna("", inplace=True)

        df["Month"] = pd.to_datetime(df["Date"]).dt.strftime("%B")
        filtered_df = df[df["Month"].str.lower() == month_name.lower()]

        if filtered_df.empty:
            return f"No records found for {month_name}."

        table_rows = []
        for _, row in filtered_df.iterrows():
            date = row["Date"]
            day = row["Day"]
            status = row["Status"]
            remarks = row["Remarks"]
            table_rows.append(f"{date} ({day}) - {status} | {remarks}")

        return "\n".join(table_rows)
    except Exception as e:
        return f"❌ Failed to read {month_name} timesheet: {e}"

# import pandas as pd
# import streamlit as st

# def fetch_timesheet_df():
#     wb = openpyxl.load_workbook("attendance.xlsx")
#     ws = wb.active
#     data = []
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         date_value, day, status, remarks = row
#         date_str = date_value.strftime("%Y-%m-%d") if isinstance(date_value, datetime) else str(date_value)
#         data.append({
#             "Date": date_str,
#             "Day": day,
#             "Status": status if status else "",
#             "Remarks": remarks if remarks else ""
#         })
#     return pd.DataFrame(data)



def ensure_workbook_exists():
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(("Date", "Day", "Status", "Remarks"))
        workbook.save(EXCEL_FILE)

def check_existing_entry(date):
    ensure_workbook_exists()
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(date):
            return {
                "date": row[0],
                "day": row[1],
                "status": row[2],
                "remarks": row[3]
            }
    return None

import pandas as pd
import os

def store_entry_to_excel(date, day, status, remarks, overwrite=False, combine=True):
    file_path = "attendance.xlsx"

    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
    else:
        df = pd.DataFrame(columns=["Date", "Day", "Status", "Remarks"])

    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    target_date = pd.to_datetime(date).date()

    existing_index = df[df["Date"] == target_date].index

    if not existing_index.empty:
        idx = existing_index[0]

        if overwrite:
            df.loc[idx] = [target_date, day, status, remarks]
        elif combine:
            old_remarks = str(df.loc[idx, "Remarks"])
            if remarks and remarks not in old_remarks:
                combined_remarks = old_remarks + "; " + remarks if old_remarks else remarks
                df.loc[idx, "Remarks"] = combined_remarks
            df.loc[idx, "Status"] = status or df.loc[idx, "Status"]
        else:
            # Do nothing if neither overwrite nor combine
            pass
    else:
        new_entry = pd.DataFrame([{
            "Date": target_date,
            "Day": day,
            "Status": status,
            "Remarks": remarks
        }])
        df = pd.concat([df, new_entry], ignore_index=True)

    df = df.sort_values(by="Date", ascending=True)
    df.to_excel(file_path, index=False)


from datetime import datetime

def fetch_timesheet():
    if not os.path.exists("attendance.xlsx"):
        return "No attendance data found."

    wb = openpyxl.load_workbook("attendance.xlsx")
    ws = wb.active

    entries = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_value, day, status, remarks = row
        date_str = datetime.strftime(date_value, "%Y-%m-%d") if isinstance(date_value, datetime) else str(date_value)
        status = status if status else "None"
        remarks = remarks if remarks else "None"
        entries.append(f"{date_str} ({day}) - {status} | {remarks}")

    return "\n".join(entries)

